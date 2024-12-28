import json
import openpyxl
import re

# Hàm để loại bỏ thẻ <b> và </b> khỏi chuỗi
def remove_b_tags(text):
    return re.sub(r'</?b>', '', text)

# Đọc dữ liệu từ file data.json
with open('data.json', 'r', encoding='utf-8') as file:
    data = json.load(file)

# Tạo hoặc mở file Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Thêm tiêu đề các cột ở row 8
ws['A8'] = ""
ws['B8'] = "Question - max 120 characters"
ws['C8'] = "Answer 1 - max 75 characters"
ws['D8'] = "Answer 2 - max 75 characters"
ws['E8'] = "Answer 3 - max 75 characters"
ws['F8'] = "Answer 4 - max 75 characters"
ws['G8'] = "Time limit (sec) – 5, 10, 20, 30, 60, 90, 120, or 240 secs"
ws['H8'] = "Correct answer(s) - choose at least one"

# Bắt đầu từ row 9
row_start = 9
for index, question_data in enumerate(data['quiz'], start=row_start):
    # Loại bỏ thẻ <b> trong nội dung câu hỏi
    question = remove_b_tags(question_data['question'])
    answer_A = remove_b_tags(question_data['a'])
    answer_B = remove_b_tags(question_data['b'])
    answer_C = remove_b_tags(question_data['c'])
    answer_D = remove_b_tags(question_data['d'])
    
    # Quy đổi đáp án từ chữ cái sang số (A -> 1, B -> 2, C -> 3, D -> 4)
    answer_mapping = {'a': 1, 'b': 2, 'c': 3, 'd': 4}
    correct_answer = answer_mapping[question_data['correct']]

    # Điền dữ liệu vào các ô tương ứng
    ws[f'A{index}'] = index - row_start + 1
    ws[f'B{index}'] = question
    ws[f'C{index}'] = answer_A
    ws[f'D{index}'] = answer_B
    ws[f'E{index}'] = answer_C
    ws[f'F{index}'] = answer_D
    ws[f'G{index}'] = 30  # Thời gian trả lời mặc định là 30
    ws[f'H{index}'] = correct_answer  # Điền đáp án

# Lưu file Excel
wb.save('data.xlsx')
print("Đã hoàn thành việc điền dữ liệu vào file Excel.")
